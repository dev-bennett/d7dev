"""
Retention Sizing Model — Excel Workbook Builder (v3)
Single-sheet compact layout with all inputs and outputs visible together.
Email funnel parameters are a single set across all plans.
Enterprise and New Subscriber segments excluded.

Funnel:
  Segment Size × Delivery × Open × Click = Re-engaged (Clicked)
  Clicked × Retention Rate = Retained (Gross)
  Size × Delivery × Open × Reminder Churn Rate = Reminder Churn Loss
  Net = Retained - Reminder Churn
  Revenue = Net × ARPU × Remaining Months
"""

import csv
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
V1_DIR = BASE_DIR.parent / 'v1_proposal'

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
INPUT_FILL = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
CALC_FILL = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
OUTPUT_FILL = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
NEGATIVE_FILL = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
SECTION_BG = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
LABEL_FONT = Font(name='Calibri', size=11)
BOLD_FONT = Font(name='Calibri', size=11, bold=True)
TITLE_FONT = Font(name='Calibri', size=14, bold=True)
SECTION_FONT = Font(name='Calibri', size=12, bold=True, color='374151')
NOTE_FONT = Font(name='Calibri', size=9, italic=True, color='6B7280')
THIN_BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)
PCT_FMT = '0.0%'
PCT2_FMT = '0.00%'
NUM_FMT = '#,##0'
NUM1_FMT = '#,##0.0'
DOLLAR_FMT = '$#,##0.00'
DOLLAR_WHOLE = '$#,##0'


def style_header_row(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER


def set_cell(ws, row, col, value=None, font=None, fill=None, fmt=None, bold=False):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    elif bold:
        cell.font = BOLD_FONT
    else:
        cell.font = LABEL_FONT
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    cell.border = THIN_BORDER
    return cell


# ---------------------------------------------------------------------------
# Load data
# ---------------------------------------------------------------------------
CORE_PLANS = ['business', 'creator', 'pro', 'pro-plus']
SEGMENTS = ['ACTIVE_DOWNLOADER', 'ACTIVE_BROWSER', 'EARLY_LAPSE', 'DEEP_LAPSE', 'DORMANT']
SEGMENT_LABELS = {
    'ACTIVE_DOWNLOADER': 'Active Downloader',
    'ACTIVE_BROWSER': 'Active Browser',
    'EARLY_LAPSE': 'Early Lapse',
    'DEEP_LAPSE': 'Deep Lapse',
    'DORMANT': 'Dormant',
}
PLAN_LABELS = {
    'business': 'Business',
    'creator': 'Personal (Creator)',
    'pro': 'Pro',
    'pro-plus': 'Pro Plus',
}

# S1: 6-month trailing averages (Sep 2025 – Feb 2026)
TARGET_MONTHS = {'2025-09', '2025-10', '2025-11', '2025-12', '2026-01', '2026-02'}
seg_totals = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

with open(V1_DIR / 's1.csv') as f:
    for row in csv.DictReader(f):
        plan = row['PLAN_TYPE']
        if plan not in CORE_PLANS:
            continue
        month = row['MONTH_START'][:7]
        if month not in TARGET_MONTHS:
            continue
        seg = row['LIFECYCLE_SEGMENT']
        if seg not in SEGMENTS:
            continue
        seg_totals[plan][seg][month] = int(row['SUBSCRIBER_COUNT'])

seg_avg = {}
for plan in CORE_PLANS:
    for seg in SEGMENTS:
        values = [seg_totals[plan][seg].get(m, 0) for m in TARGET_MONTHS]
        seg_avg[plan, seg] = round(sum(values) / len(values))

# R1: ARPU
arpu = {}
with open(BASE_DIR / 'r1.csv') as f:
    for row in csv.DictReader(f):
        if row['PLAN_TYPE'] in CORE_PLANS:
            arpu[row['PLAN_TYPE']] = float(row['AVG_MONTHLY_REVENUE'])

# R2: Churn + tenure
churn = {}
remaining = {}
with open(BASE_DIR / 'r2.csv') as f:
    for row in csv.DictReader(f):
        if row['PLAN_TYPE'] in CORE_PLANS:
            churn[row['PLAN_TYPE']] = float(row['AVG_MONTHLY_CHURN_RATE_PCT']) / 100
            remaining[row['PLAN_TYPE']] = float(row['IMPLIED_REMAINING_MONTHS'])

# Default parameters — single set across all plans
DEFAULT_PARAMS = {
    'ACTIVE_DOWNLOADER': {'delivery': 0.95, 'open': 0.25, 'click': 0.08, 'retention': 0.15, 'reminder_churn': 0.00},
    'ACTIVE_BROWSER':    {'delivery': 0.95, 'open': 0.28, 'click': 0.12, 'retention': 0.20, 'reminder_churn': 0.00},
    'EARLY_LAPSE':       {'delivery': 0.95, 'open': 0.22, 'click': 0.08, 'retention': 0.20, 'reminder_churn': 0.00},
    'DEEP_LAPSE':        {'delivery': 0.93, 'open': 0.15, 'click': 0.05, 'retention': 0.10, 'reminder_churn': 0.05},
    'DORMANT':           {'delivery': 0.90, 'open': 0.10, 'click': 0.03, 'retention': 0.05, 'reminder_churn': 0.10},
}


# ---------------------------------------------------------------------------
# Build workbook — single sheet
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = 'Retention Model'
ws.sheet_properties.tabColor = '059669'

# ── Title ──
ws.merge_cells('A1:J1')
set_cell(ws, 1, 1, 'Lifecycle Email Retention Model — Initial Ramp-Up', font=TITLE_FONT)
ws.merge_cells('A2:J2')
set_cell(ws, 2, 1,
         'Blue cells = adjustable inputs. Gray = calculated. Green = key outputs. Red = negative effects.',
         font=NOTE_FONT)

r = 4  # current row tracker

# =====================================================================
# SECTION A: Plan Inputs (ARPU & Churn)
# =====================================================================
set_cell(ws, r, 1, 'Plan Inputs', font=SECTION_FONT)
r += 1
for c, h in enumerate(['Plan Type', 'Monthly ARPU', 'Monthly Churn', 'Remaining Months'], 1):
    ws.cell(r, c, h)
style_header_row(ws, r, 1, 4)

PLAN_START = r + 1
for i, plan in enumerate(CORE_PLANS):
    row = PLAN_START + i
    set_cell(ws, row, 1, PLAN_LABELS[plan])
    set_cell(ws, row, 2, arpu[plan], fill=INPUT_FILL, fmt=DOLLAR_FMT)
    set_cell(ws, row, 3, churn[plan], fill=INPUT_FILL, fmt=PCT2_FMT)
    set_cell(ws, row, 4, f'=MIN(60, 1/C{row})', fill=CALC_FILL, fmt='0.0')
PLAN_END = PLAN_START + len(CORE_PLANS) - 1

r = PLAN_END + 2

# =====================================================================
# SECTION B: Email Funnel Parameters (single set, all plans)
# =====================================================================
set_cell(ws, r, 1, 'Email Funnel Parameters', font=SECTION_FONT)
r += 1
param_headers = ['Segment', 'Delivery Rate', 'Open Rate', 'Click Rate', 'Retention Rate', 'Reminder Churn']
for c, h in enumerate(param_headers, 1):
    ws.cell(r, c, h)
style_header_row(ws, r, 1, len(param_headers))

PARAM_START = r + 1
param_keys = ['delivery', 'open', 'click', 'retention', 'reminder_churn']
for i, seg in enumerate(SEGMENTS):
    row = PARAM_START + i
    set_cell(ws, row, 1, SEGMENT_LABELS[seg])
    for j, key in enumerate(param_keys):
        set_cell(ws, row, 2 + j, DEFAULT_PARAMS[seg][key], fill=INPUT_FILL, fmt=PCT_FMT)
PARAM_END = PARAM_START + len(SEGMENTS) - 1

r = PARAM_END + 2

# =====================================================================
# SECTION C: Segment Sizes (reference data)
# =====================================================================
set_cell(ws, r, 1, 'Segment Sizes (6-Month Avg, excl. Enterprise & New)', font=SECTION_FONT)
r += 1
size_headers = ['Segment'] + [PLAN_LABELS[p] for p in CORE_PLANS] + ['Total']
for c, h in enumerate(size_headers, 1):
    ws.cell(r, c, h)
style_header_row(ws, r, 1, len(size_headers))

SIZE_START = r + 1
for i, seg in enumerate(SEGMENTS):
    row = SIZE_START + i
    set_cell(ws, row, 1, SEGMENT_LABELS[seg])
    total = 0
    for j, plan in enumerate(CORE_PLANS):
        val = seg_avg[plan, seg]
        total += val
        set_cell(ws, row, 2 + j, val, fmt=NUM_FMT)
    set_cell(ws, row, 2 + len(CORE_PLANS), total, fmt=NUM_FMT, bold=True)
SIZE_END = SIZE_START + len(SEGMENTS) - 1

# Total row
size_total_r = SIZE_END + 1
set_cell(ws, size_total_r, 1, 'Total', bold=True)
for j in range(len(CORE_PLANS) + 1):
    col = 2 + j
    cl = get_column_letter(col)
    set_cell(ws, size_total_r, col, f'=SUM({cl}{SIZE_START}:{cl}{SIZE_END})', fmt=NUM_FMT, bold=True)

r = size_total_r + 2

# =====================================================================
# SECTION D: Funnel Output — compact matrix
# One row per segment, columns grouped by plan showing key outputs
# =====================================================================
set_cell(ws, r, 1, 'Funnel Output (Monthly)', font=SECTION_FONT)
r += 1

# Build headers: Segment | [Plan: Size, Net Retained, Rev Saved, LTV Saved] x4 | Totals
out_headers = ['Segment']
plan_col_start = {}  # plan -> starting column for that plan's group
col = 2
for plan in CORE_PLANS:
    plan_col_start[plan] = col
    out_headers.extend([
        f'{PLAN_LABELS[plan]}',    # Size
        'Net/Mo',                   # Net retained
        'Rev/Mo',                   # Monthly rev saved
    ])
    col += 3
# Total columns
total_col_start = col
out_headers.extend(['Total Net/Mo', 'Total Rev/Mo', 'Total LTV/Mo'])

# Merge plan name headers across their columns
header_row = r
for c, h in enumerate(out_headers, 1):
    ws.cell(header_row, c, h)
style_header_row(ws, header_row, 1, len(out_headers))

# Merge plan labels across their 3 sub-columns
plan_label_row = r - 1  # use the row above for plan group labels
# Actually, let's put plan group labels in a row above the headers
# Shift: insert a plan-group row
# We need to be more careful. Let me use a two-row header approach.

# Overwrite: use a simpler flat header approach
# Segment | Biz Size | Biz Net | Biz Rev | Cre Size | Cre Net | Cre Rev | ... | Tot Net | Tot Rev | Tot LTV
# This is already what we have. Let's refine the labels.

# Re-do headers with short plan abbreviations
PLAN_SHORT = {
    'business': 'Biz',
    'creator': 'Creator',
    'pro': 'Pro',
    'pro-plus': 'Pro+',
}

# Clear and redo
out_headers = ['Segment']
col = 2
for plan in CORE_PLANS:
    plan_col_start[plan] = col
    short = PLAN_SHORT[plan]
    out_headers.extend([f'{short} Size', f'{short} Net', f'{short} Rev'])
    col += 3
total_col_start = col
out_headers.extend(['All Net/Mo', 'All Rev/Mo', 'All LTV'])

for c, h in enumerate(out_headers, 1):
    ws.cell(header_row, c, h)
style_header_row(ws, header_row, 1, len(out_headers))

OUT_START = header_row + 1
for i, seg in enumerate(SEGMENTS):
    row = OUT_START + i
    param_r = PARAM_START + i  # row in Section B for this segment's funnel params
    # Param columns: B=delivery, C=open, D=click, E=retention, F=reminder_churn

    set_cell(ws, row, 1, SEGMENT_LABELS[seg])

    net_refs = []
    rev_refs = []

    for plan_idx, plan in enumerate(CORE_PLANS):
        pc = plan_col_start[plan]  # first column for this plan
        size_col_letter = get_column_letter(2 + plan_idx)  # column in Section C
        plan_row = PLAN_START + plan_idx  # row in Section A for ARPU/churn

        # Size (reference from Section C)
        size_ref = f'{size_col_letter}{SIZE_START + i}'
        set_cell(ws, row, pc, f'={size_ref}', fmt=NUM_FMT)

        # Net Retained/Mo = Size × Delivery × Open × (Click × Retention - ReminderChurn)
        # Broken out:
        #   Retained = Size × Delivery × Open × Click × Retention
        #   ReminderChurn = Size × Delivery × Open × ReminderChurn
        #   Net = Retained - ReminderChurn
        size_cell = f'{get_column_letter(pc)}{row}'
        net_formula = (
            f'={size_cell}'
            f'*$B${param_r}'   # delivery
            f'*$C${param_r}'   # open
            f'*($D${param_r}*$E${param_r}'   # click × retention
            f'-$F${param_r})'  # minus reminder churn
        )
        net_col = pc + 1
        set_cell(ws, row, net_col, net_formula, fill=OUTPUT_FILL, fmt=NUM1_FMT)

        # Rev Saved/Mo = Net × ARPU
        net_cell = f'{get_column_letter(net_col)}{row}'
        rev_formula = f'={net_cell}*$B${plan_row}'
        rev_col = pc + 2
        set_cell(ws, row, rev_col, rev_formula, fill=OUTPUT_FILL, fmt=DOLLAR_WHOLE)

        net_refs.append(f'{get_column_letter(net_col)}{row}')
        rev_refs.append(f'{get_column_letter(rev_col)}{row}')

    # Total columns
    tc = total_col_start
    # All Net/Mo
    set_cell(ws, row, tc, f'={"+".join(net_refs)}', fill=OUTPUT_FILL, fmt=NUM1_FMT, bold=True)
    # All Rev/Mo
    set_cell(ws, row, tc + 1, f'={"+".join(rev_refs)}', fill=OUTPUT_FILL, fmt=DOLLAR_WHOLE, bold=True)
    # All LTV (sum of each plan's rev × remaining months)
    ltv_parts = []
    for plan_idx, plan in enumerate(CORE_PLANS):
        rev_col = plan_col_start[plan] + 2
        plan_row = PLAN_START + plan_idx
        ltv_parts.append(f'{get_column_letter(rev_col)}{row}*$D${plan_row}')
    set_cell(ws, row, tc + 2, f'={"+".join(ltv_parts)}', fill=OUTPUT_FILL, fmt=DOLLAR_WHOLE, bold=True)

OUT_END = OUT_START + len(SEGMENTS) - 1

# Total row
out_total_r = OUT_END + 1
set_cell(ws, out_total_r, 1, 'TOTAL', bold=True)
for col_idx in range(2, len(out_headers) + 1):
    cl = get_column_letter(col_idx)
    fmt = NUM_FMT if any(f'{PLAN_SHORT[p]} Size' == out_headers[col_idx - 1] for p in CORE_PLANS) else (
        DOLLAR_WHOLE if 'Rev' in out_headers[col_idx - 1] or 'LTV' in out_headers[col_idx - 1] else NUM1_FMT
    )
    fill = OUTPUT_FILL if 'Net' in out_headers[col_idx - 1] or 'Rev' in out_headers[col_idx - 1] or 'LTV' in out_headers[col_idx - 1] else CALC_FILL
    set_cell(ws, out_total_r, col_idx,
             f'=SUM({cl}{OUT_START}:{cl}{OUT_END})',
             bold=True, fmt=fmt, fill=fill)

r = out_total_r + 2

# =====================================================================
# SECTION E: Summary
# =====================================================================
set_cell(ws, r, 1, 'Summary by Plan Type', font=SECTION_FONT)
r += 1
for c, h in enumerate(['Plan Type', 'Net Retained/Mo', 'Monthly Rev Saved', 'LTV Saved'], 1):
    ws.cell(r, c, h)
style_header_row(ws, r, 1, 4)

SUM_START = r + 1
for i, plan in enumerate(CORE_PLANS):
    row = SUM_START + i
    plan_row = PLAN_START + i
    pc = plan_col_start[plan]
    net_col = pc + 1
    rev_col = pc + 2
    net_cl = get_column_letter(net_col)
    rev_cl = get_column_letter(rev_col)

    set_cell(ws, row, 1, PLAN_LABELS[plan])
    # Sum net across segments for this plan
    set_cell(ws, row, 2, f'=SUM({net_cl}{OUT_START}:{net_cl}{OUT_END})',
             fill=OUTPUT_FILL, fmt=NUM1_FMT, bold=True)
    # Sum rev across segments for this plan
    set_cell(ws, row, 3, f'=SUM({rev_cl}{OUT_START}:{rev_cl}{OUT_END})',
             fill=OUTPUT_FILL, fmt=DOLLAR_WHOLE)
    # LTV = monthly rev × remaining months
    set_cell(ws, row, 4, f'=C{row}*$D${plan_row}',
             fill=OUTPUT_FILL, fmt=DOLLAR_WHOLE)
SUM_END = SUM_START + len(CORE_PLANS) - 1

# Grand total
grand_r = SUM_END + 1
set_cell(ws, grand_r, 1, 'GRAND TOTAL', bold=True)
for col in range(2, 5):
    cl = get_column_letter(col)
    set_cell(ws, grand_r, col,
             f'=SUM({cl}{SUM_START}:{cl}{SUM_END})',
             bold=True, fmt=NUM1_FMT if col == 2 else DOLLAR_WHOLE,
             fill=OUTPUT_FILL)

r = grand_r + 2

# ── How to Read ──
set_cell(ws, r, 1, 'How to Read This Model', font=SECTION_FONT)
notes = [
    'Net Retained/Mo: subscribers saved from churning minus reminder churn. Monthly subscriber impact.',
    'Rev Saved/Mo: net retained × plan ARPU. Monthly incremental revenue from the program.',
    'LTV Saved: monthly rev × implied remaining months. Total lifetime value protected.',
    'To adjust: change the blue cells in Plan Inputs or Email Funnel Parameters. All outputs update.',
    'Reminder Churn: Deep Lapse/Dormant subscribers who open an email and cancel (reminded they\'re paying). Set to 0% for active segments.',
    'Retention Rate: of subscribers who click, what % are saved from churning. Core value driver.',
    'Segment sizes are 6-month trailing averages (Sep 2025 - Feb 2026). Enterprise and new subscribers excluded.',
    'Default funnel parameters are starting points for discussion, not calibrated predictions.',
]
for i, n in enumerate(notes):
    ws.merge_cells(f'A{r+1+i}:J{r+1+i}')
    set_cell(ws, r + 1 + i, 1, f'{i+1}. {n}', font=NOTE_FONT)

# ── Column widths ──
ws.column_dimensions['A'].width = 22
for c in range(2, len(out_headers) + 1):
    header_text = out_headers[c - 1] if c - 1 < len(out_headers) else ''
    if 'Size' in header_text:
        ws.column_dimensions[get_column_letter(c)].width = 12
    elif 'Net' in header_text or 'Rev' in header_text or 'LTV' in header_text:
        ws.column_dimensions[get_column_letter(c)].width = 14
    else:
        ws.column_dimensions[get_column_letter(c)].width = 16

# Save
output_path = BASE_DIR / 'retention_model.xlsx'
wb.save(output_path)
print(f'Generated: {output_path}')
