"""
Evergreen Lifecycle Email Retention Model — Excel Workbook Builder
Single-sheet compact layout.

Key difference from ramp-up model:
  - Denominator is monthly INFLOW into each segment (transition volume),
    not static segment size
  - Accounts for early exits (subscribers who re-engage before series completes)
  - Accounts for cooldown blocking (subscribers who re-enter a segment within
    cooldown period are not re-enrolled)
  - Accounts for deduplication (subscribers in a higher-priority flow are excluded)

Funnel per flow:
  Monthly Inflow × (1 - Cooldown Block Rate) × (1 - Dedup Block Rate)
    = Effective Enrollments
  Effective Enrollments × Avg Emails Received / Series Length × Delivery × Open × Click
    = Clicked (re-engaged via email)
  Clicked × Retention Rate = Retained (Gross)
  Effective Enrollments × Avg Emails Received / Series Length × Delivery × Open × Reminder Churn
    = Reminder Churn Loss
  Net = Retained - Reminder Churn
  Revenue = Net × Weighted ARPU × Weighted Remaining Months
"""

import csv
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
V1_DIR = BASE_DIR.parent / 'v1_proposal'
RET_DIR = BASE_DIR.parent / 'retention_sizing'

# ---------------------------------------------------------------------------
# Styles (same as ramp-up model for consistency)
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
INPUT_FILL = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
CALC_FILL = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
OUTPUT_FILL = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
NEGATIVE_FILL = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
LABEL_FONT = Font(name='Calibri', size=11)
BOLD_FONT = Font(name='Calibri', size=11, bold=True)
TITLE_FONT = Font(name='Calibri', size=14, bold=True)
SECTION_FONT = Font(name='Calibri', size=12, bold=True, color='374151')
NOTE_FONT = Font(name='Calibri', size=9, italic=True, color='6B7280')
THIN_BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)
PCT_FMT = '0.0%'
NUM_FMT = '#,##0'
NUM1_FMT = '#,##0.0'
DOLLAR_WHOLE = '$#,##0'
DOLLAR_FMT = '$#,##0.00'
PCT2_FMT = '0.00%'


def style_header_row(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER


def set_cell(ws, row, col, value=None, font=None, fill=None, fmt=None, bold=False):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    elif bold:
        cell.font = BOLD_FONT
    else:
        cell.font = LABEL_FONT
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    cell.border = THIN_BORDER
    return cell


# ---------------------------------------------------------------------------
# Load data
# ---------------------------------------------------------------------------
CORE_PLANS = ['business', 'creator', 'pro', 'pro-plus']

# R1: ARPU
arpu = {}
with open(RET_DIR / 'r1.csv') as f:
    for row in csv.DictReader(f):
        if row['PLAN_TYPE'] in CORE_PLANS:
            arpu[row['PLAN_TYPE']] = float(row['AVG_MONTHLY_REVENUE'])

# R2: Churn + remaining months
churn = {}
remaining = {}
with open(RET_DIR / 'r2.csv') as f:
    for row in csv.DictReader(f):
        if row['PLAN_TYPE'] in CORE_PLANS:
            churn[row['PLAN_TYPE']] = float(row['AVG_MONTHLY_CHURN_RATE_PCT']) / 100
            remaining[row['PLAN_TYPE']] = float(row['IMPLIED_REMAINING_MONTHS'])

# Weighted ARPU and remaining months across plans (weighted by segment inflow)
# We'll compute this from S2 transition data
# S2: Transition matrix
transitions = defaultdict(int)  # (prior, current) -> avg monthly
plan_transitions = defaultdict(lambda: defaultdict(int))  # plan -> (prior, current) -> avg monthly

with open(V1_DIR / 's2.csv') as f:
    for row in csv.DictReader(f):
        plan = row['PLAN_TYPE']
        if plan not in CORE_PLANS:
            continue
        prior = row['PRIOR_SEGMENT']
        current = row['CURRENT_SEGMENT']
        avg_mo = int(row['AVG_MONTHLY_TRANSITIONS'])
        transitions[(prior, current)] += avg_mo
        plan_transitions[plan][(prior, current)] = avg_mo

# ---------------------------------------------------------------------------
# Compute flow enrollment volumes from transition data
# ---------------------------------------------------------------------------

# Flow definitions: which transitions trigger enrollment
FLOWS = [
    {
        'name': 'Early Lapse Re-engagement',
        'short': 'EL Re-engage',
        'target_segment': 'EARLY_LAPSE',
        'inflow_from': ['ACTIVE_DOWNLOADER', 'ACTIVE_BROWSER'],
        'series_length': 3,
        'cooldown_days': 60,
        'series_days': 21,
    },
    {
        'name': 'Deep Lapse Win-back',
        'short': 'DL Win-back',
        'target_segment': 'DEEP_LAPSE',
        'inflow_from': ['EARLY_LAPSE', 'ACTIVE_DOWNLOADER', 'ACTIVE_BROWSER'],
        'series_length': 2,
        'cooldown_days': 90,
        'series_days': 30,
    },
    {
        'name': 'Dormant Sunset',
        'short': 'Dormant',
        'target_segment': 'DORMANT',
        'inflow_from': ['DEEP_LAPSE'],
        'series_length': 2,
        'cooldown_days': None,  # terminal
        'series_days': 14,
    },
    {
        'name': 'Active Browser Download Nudge',
        'short': 'AB Nudge',
        'target_segment': 'ACTIVE_BROWSER',
        'inflow_from': None,  # stability-based, not transition-based
        'series_length': 2,
        'cooldown_days': 30,
        'series_days': 14,
        'stability_based': True,
    },
    {
        'name': 'Active Downloader Reinforcement',
        'short': 'AD Reinforce',
        'target_segment': 'ACTIVE_DOWNLOADER',
        'inflow_from': None,  # recurring monthly
        'series_length': 1,
        'cooldown_days': 30,
        'series_days': 30,
        'stability_based': True,
    },
]

# Compute monthly inflows per flow
for flow in FLOWS:
    if flow.get('stability_based'):
        # Use self-retention as proxy for stable subscribers
        seg = flow['target_segment']
        flow['monthly_inflow'] = transitions.get((seg, seg), 0)
        # Plan-weighted ARPU
        total_vol = 0
        weighted_arpu = 0
        weighted_remaining = 0
        for plan in CORE_PLANS:
            vol = plan_transitions[plan].get((seg, seg), 0)
            total_vol += vol
            weighted_arpu += vol * arpu[plan]
            weighted_remaining += vol * remaining[plan]
        flow['weighted_arpu'] = weighted_arpu / total_vol if total_vol > 0 else 0
        flow['weighted_remaining'] = weighted_remaining / total_vol if total_vol > 0 else 0
    else:
        total_inflow = 0
        weighted_arpu = 0
        weighted_remaining = 0
        for src in flow['inflow_from']:
            vol = transitions.get((src, flow['target_segment']), 0)
            total_inflow += vol
            # Weight by plan mix
            for plan in CORE_PLANS:
                pvol = plan_transitions[plan].get((src, flow['target_segment']), 0)
                weighted_arpu += pvol * arpu[plan]
                weighted_remaining += pvol * remaining[plan]
        flow['monthly_inflow'] = total_inflow
        flow['weighted_arpu'] = weighted_arpu / total_inflow if total_inflow > 0 else 0
        flow['weighted_remaining'] = weighted_remaining / total_inflow if total_inflow > 0 else 0

# Default funnel parameters per flow
DEFAULT_PARAMS = {
    'Early Lapse Re-engagement':      {'delivery': 0.95, 'open': 0.22, 'click': 0.08, 'retention': 0.20, 'reminder_churn': 0.00,
                                        'cooldown_block': 0.10, 'dedup_block': 0.00, 'early_exit_rate': 0.40},
    'Deep Lapse Win-back':            {'delivery': 0.93, 'open': 0.15, 'click': 0.05, 'retention': 0.10, 'reminder_churn': 0.05,
                                        'cooldown_block': 0.05, 'dedup_block': 0.00, 'early_exit_rate': 0.15},
    'Dormant Sunset':                 {'delivery': 0.90, 'open': 0.10, 'click': 0.03, 'retention': 0.05, 'reminder_churn': 0.10,
                                        'cooldown_block': 0.00, 'dedup_block': 0.00, 'early_exit_rate': 0.05},
    'Active Browser Download Nudge':  {'delivery': 0.95, 'open': 0.28, 'click': 0.12, 'retention': 0.20, 'reminder_churn': 0.00,
                                        'cooldown_block': 0.15, 'dedup_block': 0.05, 'early_exit_rate': 0.20},
    'Active Downloader Reinforcement':{'delivery': 0.95, 'open': 0.25, 'click': 0.08, 'retention': 0.15, 'reminder_churn': 0.00,
                                        'cooldown_block': 0.00, 'dedup_block': 0.05, 'early_exit_rate': 0.10},
}


# ---------------------------------------------------------------------------
# Build workbook — single sheet
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = 'Evergreen Model'
ws.sheet_properties.tabColor = '7c3aed'

# ── Title ──
ws.merge_cells('A1:L1')
set_cell(ws, 1, 1, 'Evergreen Lifecycle Email Retention Model', font=TITLE_FONT)
ws.merge_cells('A2:L2')
set_cell(ws, 2, 1,
         'Blue = adjustable inputs. Gray = calculated. Green = key outputs. '
         'Enrollment volumes from transition data, not static segment sizes.',
         font=NOTE_FONT)

r = 4

# =====================================================================
# SECTION A: Revenue Reference (read-only context)
# =====================================================================
set_cell(ws, r, 1, 'Revenue Reference (from ramp-up model)', font=SECTION_FONT)
r += 1
for c, h in enumerate(['Plan Type', 'Monthly ARPU', 'Monthly Churn', 'Remaining Months'], 1):
    ws.cell(r, c, h)
style_header_row(ws, r, 1, 4)

PLAN_START = r + 1
for i, plan in enumerate(CORE_PLANS):
    row = PLAN_START + i
    labels = {'business': 'Business', 'creator': 'Personal (Creator)', 'pro': 'Pro', 'pro-plus': 'Pro Plus'}
    set_cell(ws, row, 1, labels[plan])
    set_cell(ws, row, 2, arpu[plan], fill=CALC_FILL, fmt=DOLLAR_FMT)
    set_cell(ws, row, 3, churn[plan], fill=CALC_FILL, fmt=PCT2_FMT)
    set_cell(ws, row, 4, remaining[plan], fill=CALC_FILL, fmt='0.0')
PLAN_END = PLAN_START + len(CORE_PLANS) - 1

r = PLAN_END + 2

# =====================================================================
# SECTION B: Flow Parameters (adjustable)
# =====================================================================
set_cell(ws, r, 1, 'Flow Parameters', font=SECTION_FONT)
r += 1

param_headers = ['Flow', 'Monthly Inflow', 'Cooldown Block', 'Dedup Block',
                 'Early Exit Rate', 'Delivery', 'Open Rate', 'Click Rate',
                 'Retention Rate', 'Reminder Churn']
for c, h in enumerate(param_headers, 1):
    ws.cell(r, c, h)
style_header_row(ws, r, 1, len(param_headers))

FLOW_PARAM_START = r + 1
param_keys_order = ['cooldown_block', 'dedup_block', 'early_exit_rate',
                    'delivery', 'open', 'click', 'retention', 'reminder_churn']

for i, flow in enumerate(FLOWS):
    row = FLOW_PARAM_START + i
    params = DEFAULT_PARAMS[flow['name']]
    set_cell(ws, row, 1, flow['short'])
    set_cell(ws, row, 2, flow['monthly_inflow'], fill=CALC_FILL, fmt=NUM_FMT)
    for j, key in enumerate(param_keys_order):
        set_cell(ws, row, 3 + j, params[key], fill=INPUT_FILL, fmt=PCT_FMT)
FLOW_PARAM_END = FLOW_PARAM_START + len(FLOWS) - 1

r = FLOW_PARAM_END + 2

# =====================================================================
# SECTION C: Funnel Output
# =====================================================================
set_cell(ws, r, 1, 'Funnel Output (Monthly)', font=SECTION_FONT)
r += 1

out_headers = ['Flow', 'Monthly Inflow', 'Effective Enrollments',
               'Retained (Gross)', 'Reminder Churn', 'Net Retained/Mo',
               'Wtd ARPU', 'Rev Saved/Mo', 'Wtd Remaining Mo', 'LTV Saved']
for c, h in enumerate(out_headers, 1):
    ws.cell(r, c, h)
style_header_row(ws, r, 1, len(out_headers))

OUT_START = r + 1
for i, flow in enumerate(FLOWS):
    row = OUT_START + i
    pr = FLOW_PARAM_START + i  # param row

    # A: Flow name
    set_cell(ws, row, 1, flow['short'])

    # B: Monthly Inflow (reference from params)
    set_cell(ws, row, 2, f'=B{pr}', fmt=NUM_FMT)

    # C: Effective Enrollments = Inflow × (1 - Cooldown Block) × (1 - Dedup Block)
    set_cell(ws, row, 3, f'=B{row}*(1-C{pr})*(1-D{pr})', fill=CALC_FILL, fmt=NUM_FMT)

    # D: Retained (Gross)
    # = Enrollments × (1 - EarlyExitRate) × Delivery × Open × Click × Retention
    # + Enrollments × EarlyExitRate × (1/SeriesLen) × Delivery × Open × Click × Retention
    # Simplified: Enrollments × AvgSeriesCompletion × Delivery × Open × Click × Retention
    # where AvgSeriesCompletion = (1-EER) + EER×(1/SL) ≈ (1 - EER × (1 - 1/SL))
    # For simplicity, use: Enrollments × (1 - EarlyExitRate/2) × Delivery × Open × Click × Retention
    # (assumes early exiters receive ~half the series on average)
    set_cell(ws, row, 4,
             f'=C{row}*(1-E{pr}/2)*F{pr}*G{pr}*H{pr}*I{pr}',
             fill=CALC_FILL, fmt=NUM1_FMT)

    # E: Reminder Churn = Enrollments × (1 - EarlyExitRate/2) × Delivery × Open × ReminderChurn
    set_cell(ws, row, 5,
             f'=C{row}*(1-E{pr}/2)*F{pr}*G{pr}*J{pr}',
             fill=NEGATIVE_FILL, fmt=NUM1_FMT)

    # F: Net Retained/Mo = Gross - Reminder Churn
    set_cell(ws, row, 6, f'=D{row}-E{row}', fill=OUTPUT_FILL, fmt=NUM1_FMT, bold=True)

    # G: Weighted ARPU (from transition data, not adjustable)
    set_cell(ws, row, 7, flow['weighted_arpu'], fill=CALC_FILL, fmt=DOLLAR_FMT)

    # H: Rev Saved/Mo = Net × Weighted ARPU
    set_cell(ws, row, 8, f'=F{row}*G{row}', fill=OUTPUT_FILL, fmt=DOLLAR_WHOLE)

    # I: Weighted Remaining Months
    set_cell(ws, row, 9, flow['weighted_remaining'], fill=CALC_FILL, fmt='0.0')

    # J: LTV Saved = Rev/Mo × Remaining
    set_cell(ws, row, 10, f'=H{row}*I{row}', fill=OUTPUT_FILL, fmt=DOLLAR_WHOLE)

OUT_END = OUT_START + len(FLOWS) - 1

# Total row
total_r = OUT_END + 1
set_cell(ws, total_r, 1, 'TOTAL', bold=True)
for col in [2, 3]:
    cl = get_column_letter(col)
    set_cell(ws, total_r, col, f'=SUM({cl}{OUT_START}:{cl}{OUT_END})', bold=True, fmt=NUM_FMT)
for col in [4, 5, 6]:
    cl = get_column_letter(col)
    set_cell(ws, total_r, col, f'=SUM({cl}{OUT_START}:{cl}{OUT_END})', bold=True, fmt=NUM1_FMT,
             fill=OUTPUT_FILL if col == 6 else (NEGATIVE_FILL if col == 5 else CALC_FILL))
for col in [8, 10]:
    cl = get_column_letter(col)
    set_cell(ws, total_r, col, f'=SUM({cl}{OUT_START}:{cl}{OUT_END})', bold=True, fmt=DOLLAR_WHOLE,
             fill=OUTPUT_FILL)

r = total_r + 2

# =====================================================================
# SECTION D: Notes
# =====================================================================
set_cell(ws, r, 1, 'How to Read This Model', font=SECTION_FONT)
notes = [
    'Monthly Inflow: subscribers transitioning INTO each segment per month (from s2 transition data).',
    'Cooldown Block: % of inflow blocked because subscriber completed this flow within the cooldown period.',
    'Dedup Block: % of inflow blocked because subscriber is already in a higher-priority flow.',
    'Early Exit Rate: % of enrolled subscribers who re-engage (or escalate) before completing the series.',
    'Early exiters are modeled as receiving ~half the series on average.',
    'Retention Rate: of subscribers who click through, what % are saved from churning.',
    'Reminder Churn: of subscribers who open, what % cancel (Dormant/Deep Lapse only).',
    'Weighted ARPU and Remaining Months are plan-weighted by each flow\'s transition mix (not adjustable here).',
    'All blue cells are adjustable. Cooldown/dedup/early-exit rates are estimates — calibrate after deployment.',
    'Inflow volumes will update when s2.csv is refreshed with rolling-window query data.',
]
for i, n in enumerate(notes):
    ws.merge_cells(f'A{r+1+i}:L{r+1+i}')
    set_cell(ws, r + 1 + i, 1, f'{i+1}. {n}', font=NOTE_FONT)

# ── Column widths ──
ws.column_dimensions['A'].width = 24
for c in range(2, len(out_headers) + 1):
    ws.column_dimensions[get_column_letter(c)].width = 16

# Save
output_path = BASE_DIR / 'evergreen_model.xlsx'
wb.save(output_path)
print(f'Generated: {output_path}')
